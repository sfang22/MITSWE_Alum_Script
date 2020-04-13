from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from openpyxl import Workbook
kerb = ""
pwd = ""

wb = Workbook()
ws = wb.active
ws.title = "swe alum info"
ws['A1'] = "Name"
ws['B1'] = "Email"
wb.save("info.xlsx")



driver_loc = './chromedriver'
driver = webdriver.Chrome(driver_loc)
driver.get("https://alum.mit.edu/directory/#/directory-search-results")
username = driver.find_element_by_id("username")
password = driver.find_element_by_id("password")

#login
username.send_keys(kerb)
password.send_keys(pwd)
driver.find_elements_by_class_name("button")[1].send_keys(Keys.ENTER)

#clicking on more activities
driver.find_elements_by_class_name("filter-groups__more-link")[1].click()
student_activity = driver.find_elements_by_class_name("student-activity")[0]
student_activity.click()

#selecting SWE
input = driver.find_elements_by_tag_name("input")
for i in range(len(input)):
    try:
        input[i].send_keys("Society of Women Engineers")
        if (i==16):
            input[i].send_keys(Keys.ARROW_DOWN)
            input[i].send_keys(Keys.ENTER)           
    except:
        pass

alums = set()
emails = []
titles = {"Miss", "Ms.", "Dr.", "Mr."}
def get_info():
    in_progress = True
    while in_progress:
        in_progress = False
        links = driver.find_elements_by_tag_name("a")
        for link in links:
            txt = link.text
            flag = False
            for title in titles:
                if title in txt:
                    flag = True
            
            if flag and txt not in alums:
                link.click()
                time.sleep(1)
                html_text = driver.find_element_by_tag_name("html").text
                html_text = html_text.split("\n")
                for line in html_text:
                    if "@" in line:
                        email = line
                        break
                alums.add(txt)
                emails.append(email)
                print(txt, email)
                ws['A' + str(len(alums)+1)] = txt
                ws['B' + str(len(alums)+1)] = email
                wb.save("info.xlsx")
                driver.find_elements_by_class_name("arrow-back")[0].click()
                time.sleep(1)
                in_progress = True
                break
    print(len(emails))

while True:
    try:
        get_info()
        time.sleep(1)
        driver.find_elements_by_class_name("pagination-next")[0].click()
    except:
        break

driver.quit()


